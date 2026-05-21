"""Write extracted values into a new-format P&L template workbook."""

from __future__ import annotations

import logging
import os
import re
import shutil
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

log = logging.getLogger(__name__)

ANCHOR_ROWS = [8 + 15 * i for i in range(10)]
MAX_SKU_SLOTS = len(ANCHOR_ROWS)

SKU_BLOCK_FIRST_ROW = 8
SKU_BLOCK_ROW_STEP = 15
SKU_BLOCK_LAST_ROW = 143
SKU_BLOCK_MARKET_COL = "C"
LAUNCH_YEAR_COL = "K"
LAUNCH_MONTH_COL = "L"
LAUNCH_YEAR_Q_COL = "launch_year_Q"

# First calendar month of each fiscal/calendar quarter (Actual Launch month).
QUARTER_START_MONTH: dict[int, str] = {
    1: "January",
    2: "April",
    3: "July",
    4: "October",
}

_LAUNCH_YEAR_Q_RE = re.compile(r"(20\d{2})\s*Q\s*([1-4])", re.IGNORECASE)

ANP_PROJ_COLS = [
    "Market",
    "ANP_perc of Net Sales_Year 1",
    "ANP_perc of Net Sales_Year 2",
    "ANP_perc of Net Sales_Year 3",
    "ANP_perc of Net Sales_Year 4",
    "ANP_perc of Net Sales_Year 5",
]

# Home Tab: unique Market names from extractions_proj (column order = first CSV appearance)
HOME_TAB_UNIQUE_MARKET_COL = "I"
HOME_TAB_UNIQUE_MARKET_START_ROW = 19
HOME_TAB_UNIQUE_MARKET_CLEAR_ROWS = 80

# Home Tab: base currency (template label "base currency" area)
HOME_TAB_BASE_CURRENCY_COL = "C"
HOME_TAB_BASE_CURRENCY_ROW = 6
HOME_TAB_BASE_CURRENCY_VALUE = "AUD"

FORECAST_COLS = [
    "forecast_volume_y1",
    "forecast_volume_y2",
    "forecast_volume_y3",
    "forecast_volume_y4",
    "forecast_volume_y5",
]

COGS_OTHER_COLS = [f"COGS_other_combined_y{y}" for y in range(1, 6)]
GTN_COL = "GTN_perc excluding Launch yr one-time costs"
COGS_COL = "COGS/unit"
LAUNCH_COL = "Launch Yr one-time cost (Listing fees, launch COOP)"

DED_COL_LETTERS = ["BA", "BN", "CA", "CN", "DA", "DN"]
VOL_COL_LETTERS = ["AQ", "AR", "AS", "AT", "AU"]
COGS_OTHER_COL_LETTERS = ["AK", "AL", "AM", "AN", "AO"]

# A&P ALLOC: same sku-block geometry as VOL/COGS/etc.; column X = unique Market per SKU block
AP_ALLOC_SHEET_NAME = "A&P ALLOC"
AP_ALLOC_MARKET_COL_X = "X"

# CAPEX: project-level $ amounts (not repeated per sku block); one column, consecutive rows
CAPEX_SHEET_NAME = "CAPEX"
CAPEX_PROJ_VALUE_COL = "CAPEX_CAPEX $"  # pivoted column from extract_proj (row label "CAPEX $" in source)
CAPEX_VALUE_COL_AK = "AK"
CAPEX_VALUE_START_ROW = 8
CAPEX_AK_CLEAR_ROWS = 120

# CANNIBAL: sku-block rows; year % → BA/BB/BC/BD/BE (feeds CANNIBAL_REV GA:GE via BA:BE).
CANNIBAL_SHEET_NAME = "CANNIBAL"
CANNIBAL_REV_SHEET_NAME = "CANNIBAL_REV"
CANNI_PERC_PROJ_COLS = [f"canni_perc_y{y}" for y in range(1, 6)]
CANNI_YEAR_COL_LETTERS = ["BA", "BB", "BC", "BD", "BE"]
CANNI_GP_PROJ_COL = "canni_GP_perc"
CANNI_GP_REV_COL_LETTER = "AI"  # CANNIBAL_REV!$AI$n multiplies GA:GE revenue formulas
CANNI_CLEAR_COL_LETTERS = list(CANNI_YEAR_COL_LETTERS)


def resolve_product_sku_column(df: pd.DataFrame) -> str:
    """
    Product-level SKU name (e.g. 'Blackcurrant 16'), not sku_market composite.

    Prefer sku_name.1 when pandas suffixes a duplicate sku_name column from CSV.
    """
    if "sku_name.1" in df.columns:
        return "sku_name.1"
    if "sku_market" in df.columns and "sku_name" in df.columns:
        return "sku_name"
    if "sku_name" in df.columns:
        return "sku_name"
    raise KeyError(f"No product sku column in: {list(df.columns)!r}")


def resolve_sku_column(df: pd.DataFrame) -> str:
    """Alias for product SKU column (Home Tab, sku-block grouping)."""
    return resolve_product_sku_column(df)


def _require_columns(df: pd.DataFrame, columns: Iterable[str], label: str) -> None:
    missing = [c for c in columns if c not in df.columns]
    if missing:
        raise KeyError(f"{label} missing columns: {missing!r}")


def _numeric_for_excel_cell(value: Any) -> int | float | None:
    """Coerce to int/float for openpyxl, or None if missing / non-numeric."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    n = pd.to_numeric(value, errors="coerce")
    if pd.isna(n):
        return None
    f = float(n)
    if f.is_integer():
        return int(f)
    return f


def parse_launch_year_q(value: Any) -> tuple[int | None, str | None]:
    """
    Parse ``launch_year_Q`` (e.g. ``'2026 Q1'``) → (year, month name).

    Month is the first month of the quarter (Q1→January, Q2→April, Q3→July, Q4→October).
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None, None
    match = _LAUNCH_YEAR_Q_RE.search(text)
    if not match:
        return None, None
    year = int(match.group(1))
    quarter = int(match.group(2))
    month = QUARTER_START_MONTH.get(quarter)
    return year, month


def write_sku_block_sheet(
    ws,
    extractions: pd.DataFrame,
    *,
    data_columns: list[str],
    value_col_letters: list[str],
    repeat_value: bool = False,
    sheet_label: str = "sheet",
    write_launch_dates: bool = True,
) -> int:
    """
    Write extractions on a tab using sku blocks (8, 23, 38, …) and market sub-rows.

    For each product SKU: base_row = 8 + 15 * sku_index; filter all markets for that SKU;
    write consecutive rows with column C = Market.

    - ``repeat_value=False``: one extraction column maps to one Excel column (or single pair).
    - ``repeat_value=True``: one extraction value is copied to every ``value_col_letter``.
    """
    if len(data_columns) != len(value_col_letters) and not (
        repeat_value and len(data_columns) == 1
    ):
        raise ValueError(
            f"{sheet_label}: need matching data_columns and value_col_letters, "
            f"or repeat_value=True with one data column"
        )

    sku_col = resolve_product_sku_column(extractions)
    _require_columns(extractions, [*data_columns, "Market", sku_col], "extractions")

    market_col_idx = column_index_from_string(SKU_BLOCK_MARKET_COL)
    value_col_indices = [column_index_from_string(letter) for letter in value_col_letters]
    include_launch = write_launch_dates and LAUNCH_YEAR_Q_COL in extractions.columns
    launch_year_col_idx = column_index_from_string(LAUNCH_YEAR_COL)
    launch_month_col_idx = column_index_from_string(LAUNCH_MONTH_COL)

    for row in range(SKU_BLOCK_FIRST_ROW, SKU_BLOCK_LAST_ROW + 1):
        ws.cell(row=row, column=market_col_idx, value=None)
        for col_idx in value_col_indices:
            ws.cell(row=row, column=col_idx, value=None)
        if include_launch:
            ws.cell(row=row, column=launch_year_col_idx, value=None)
            ws.cell(row=row, column=launch_month_col_idx, value=None)

    unique_skus = list(extractions[sku_col].dropna().unique())
    written = 0
    for sku_idx, sku in enumerate(unique_skus):
        if sku_idx >= MAX_SKU_SLOTS:
            log.warning(
                "%s: %d SKUs exceed %d blocks; skipping SKU %r and later",
                sheet_label,
                len(unique_skus),
                MAX_SKU_SLOTS,
                sku,
            )
            break
        base_row = SKU_BLOCK_FIRST_ROW + SKU_BLOCK_ROW_STEP * sku_idx
        sku_rows = extractions.loc[extractions[sku_col] == sku]
        for market_offset, (_, row_s) in enumerate(sku_rows.iterrows()):
            excel_row = base_row + market_offset
            if excel_row > SKU_BLOCK_LAST_ROW:
                log.warning(
                    "%s: SKU %r market rows exceed row %d; skipping remainder",
                    sheet_label,
                    sku,
                    SKU_BLOCK_LAST_ROW,
                )
                break
            ws.cell(row=excel_row, column=market_col_idx, value=row_s["Market"])
            if repeat_value:
                val = row_s[data_columns[0]]
                for col_idx in value_col_indices:
                    ws.cell(row=excel_row, column=col_idx, value=val)
            else:
                for j, col_idx in enumerate(value_col_indices):
                    ws.cell(row=excel_row, column=col_idx, value=row_s[data_columns[j]])
            if include_launch:
                year, month = parse_launch_year_q(row_s[LAUNCH_YEAR_Q_COL])
                ws.cell(row=excel_row, column=launch_year_col_idx, value=year)
                ws.cell(row=excel_row, column=launch_month_col_idx, value=month)
            written += 1
    if write_launch_dates and not include_launch:
        log.warning(
            "%s: column %r missing — skipped Actual Launch Year/Month (cols %s, %s)",
            sheet_label,
            LAUNCH_YEAR_Q_COL,
            LAUNCH_YEAR_COL,
            LAUNCH_MONTH_COL,
        )
    elif include_launch:
        log.info(
            "%s: wrote Actual Launch Year (%s) / Month (%s) from %r",
            sheet_label,
            LAUNCH_YEAR_COL,
            LAUNCH_MONTH_COL,
            LAUNCH_YEAR_Q_COL,
        )
    return written


def write_vol_auto_forecast_volumes(ws, extractions: pd.DataFrame) -> int:
    """VOL_Auto: AQ:AU = forecast_volume_y1..y5 per sku×market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=FORECAST_COLS,
        value_col_letters=VOL_COL_LETTERS,
        sheet_label="VOL_Auto",
    )


def write_sp_list_prices(ws, extractions: pd.DataFrame) -> int:
    """SP: AI = list_price_AUD per sku×market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=["list_price_AUD"],
        value_col_letters=["AI"],
        sheet_label="SP",
    )


def write_sales_ded_td_gtn(ws, extractions: pd.DataFrame) -> int:
    """SALES_DED_TD: GTN % repeated across BA/BN/CA/CN/DA/DN per market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=[GTN_COL],
        value_col_letters=DED_COL_LETTERS,
        repeat_value=True,
        sheet_label="SALES_DED_TD",
    )


def write_cogs_unit(ws, extractions: pd.DataFrame) -> int:
    """COGS: AK = COGS/unit per sku×market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=[COGS_COL],
        value_col_letters=["AK"],
        sheet_label="COGS",
    )


def write_cogs_other_combined(ws, extractions: pd.DataFrame) -> int:
    """COGS_OTHER: AK:AO = COGS_other_combined_y1..y5 per market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=COGS_OTHER_COLS,
        value_col_letters=COGS_OTHER_COL_LETTERS,
        sheet_label="COGS_OTHER",
    )


def write_sales_ded_ka_launch(ws, extractions: pd.DataFrame) -> int:
    """SALES_DED_KA: AK = launch one-time cost per market row."""
    return write_sku_block_sheet(
        ws,
        extractions,
        data_columns=[LAUNCH_COL],
        value_col_letters=["AK"],
        sheet_label="SALES_DED_KA",
    )


def write_ap_alloc_markets_column_x(ws, extractions: pd.DataFrame) -> int:
    """
    A&P ALLOC: column **X** uses the same sku-block anchors as other tabs (rows 8, 23, …).

    Per product SKU, writes **unique** ``Market`` values (first-seen order within that SKU)
    down consecutive rows: e.g. SKU 0 → X8:X…, SKU 1 → X23:X….
    """
    sku_col = resolve_product_sku_column(extractions)
    _require_columns(extractions, ["Market", sku_col], "extractions")

    x_col_idx = column_index_from_string(AP_ALLOC_MARKET_COL_X)
    for row in range(SKU_BLOCK_FIRST_ROW, SKU_BLOCK_LAST_ROW + 1):
        ws.cell(row=row, column=x_col_idx, value=None)

    unique_skus = list(extractions[sku_col].dropna().unique())
    written = 0
    for sku_idx, sku in enumerate(unique_skus):
        if sku_idx >= MAX_SKU_SLOTS:
            log.warning(
                "%s: %d SKUs exceed %d blocks; skipping SKU %r and later",
                AP_ALLOC_SHEET_NAME,
                len(unique_skus),
                MAX_SKU_SLOTS,
                sku,
            )
            break
        base_row = SKU_BLOCK_FIRST_ROW + SKU_BLOCK_ROW_STEP * sku_idx
        sku_rows = extractions.loc[extractions[sku_col] == sku]
        series = sku_rows["Market"].drop_duplicates()
        markets_list: list[Any] = []
        for m in series.tolist():
            if m is None or (isinstance(m, float) and pd.isna(m)):
                continue
            s = str(m).strip()
            if not s or s.lower() == "nan":
                continue
            markets_list.append(m)

        for market_offset, market in enumerate(markets_list):
            excel_row = base_row + market_offset
            if excel_row > SKU_BLOCK_LAST_ROW:
                log.warning(
                    "%s: SKU %r market rows exceed row %d; skipping remainder",
                    AP_ALLOC_SHEET_NAME,
                    sku,
                    SKU_BLOCK_LAST_ROW,
                )
                break
            ws.cell(row=excel_row, column=x_col_idx, value=market)
            written += 1

    log.info(
        "%s: wrote %d Market cell(s) to column %s (sku-block layout)",
        AP_ALLOC_SHEET_NAME,
        written,
        AP_ALLOC_MARKET_COL_X,
    )
    return written


def write_capex_sheet_ak_from_proj(ws, extractions_proj: pd.DataFrame) -> int:
    """
    CAPEX sheet: write ``CAPEX_CAPEX $`` from ``extractions_proj`` into column **AK**
    starting at row **8** (**AK8**, **AK9**, …) — one row per DataFrame row, **not** sku-block stepped.
    """
    if CAPEX_PROJ_VALUE_COL not in extractions_proj.columns:
        log.warning(
            "extractions_proj has no %r — skipped %s sheet column %s",
            CAPEX_PROJ_VALUE_COL,
            CAPEX_SHEET_NAME,
            CAPEX_VALUE_COL_AK,
        )
        return 0

    ak_idx = column_index_from_string(CAPEX_VALUE_COL_AK)
    for off in range(CAPEX_AK_CLEAR_ROWS):
        ws.cell(row=CAPEX_VALUE_START_ROW + off, column=ak_idx, value=None)

    written = 0
    for i, (_, row_s) in enumerate(extractions_proj.iterrows()):
        if i >= CAPEX_AK_CLEAR_ROWS:
            log.warning(
                "%s: more than %d extractions_proj rows; truncating column %s",
                CAPEX_SHEET_NAME,
                CAPEX_AK_CLEAR_ROWS,
                CAPEX_VALUE_COL_AK,
            )
            break
        excel_row = CAPEX_VALUE_START_ROW + i
        ws.cell(
            row=excel_row,
            column=ak_idx,
            value=_numeric_for_excel_cell(row_s[CAPEX_PROJ_VALUE_COL]),
        )
        written += 1

    log.info(
        "%s: wrote %d value(s) to %s%d+ from %r",
        CAPEX_SHEET_NAME,
        written,
        CAPEX_VALUE_COL_AK,
        CAPEX_VALUE_START_ROW,
        CAPEX_PROJ_VALUE_COL,
    )
    return written


def _proj_rows_by_market(extractions_proj: pd.DataFrame) -> dict[str, pd.Series]:
    """Map stripped Market name → row (Series) in extractions_proj."""
    lookup: dict[str, pd.Series] = {}
    if "Market" not in extractions_proj.columns:
        return lookup
    for _, row in extractions_proj.iterrows():
        m = row["Market"]
        if m is None or (isinstance(m, float) and pd.isna(m)):
            continue
        key = str(m).strip()
        if key:
            lookup[key] = row
    return lookup


def write_cannibal_sheet_from_proj(
    ws,
    extractions: pd.DataFrame,
    extractions_proj: pd.DataFrame,
    ws_rev=None,
) -> int:
    """
    CANNIBAL: per sku block (8, 23, …), consecutive market rows from ``extractions``;
    values from ``extractions_proj`` by ``Market``.

    - ``canni_perc_y1`` … ``y5`` → **BA, BB, BC, BD, BE**
    - ``canni_GP_perc`` → **CANNIBAL_REV!AI**
    """
    required = ["Market", *CANNI_PERC_PROJ_COLS, CANNI_GP_PROJ_COL]
    missing = [c for c in required if c not in extractions_proj.columns]
    if missing:
        log.warning(
            "extractions_proj missing %r — skipped %s sheet",
            missing,
            CANNIBAL_SHEET_NAME,
        )
        return 0

    sku_col = resolve_product_sku_column(extractions)
    _require_columns(extractions, ["Market", sku_col], "extractions")

    year_col_indices = [column_index_from_string(letter) for letter in CANNI_YEAR_COL_LETTERS]
    gp_rev_col_idx = column_index_from_string(CANNI_GP_REV_COL_LETTER)
    clear_col_indices = [column_index_from_string(letter) for letter in CANNI_CLEAR_COL_LETTERS]

    for row in range(SKU_BLOCK_FIRST_ROW, SKU_BLOCK_LAST_ROW + 1):
        for col_idx in clear_col_indices:
            ws.cell(row=row, column=col_idx, value=None)

    proj_by_market = _proj_rows_by_market(extractions_proj)
    unique_skus = list(extractions[sku_col].dropna().unique())
    written = 0

    for sku_idx, sku in enumerate(unique_skus):
        if sku_idx >= MAX_SKU_SLOTS:
            log.warning(
                "%s: %d SKUs exceed %d blocks; skipping SKU %r and later",
                CANNIBAL_SHEET_NAME,
                len(unique_skus),
                MAX_SKU_SLOTS,
                sku,
            )
            break
        base_row = SKU_BLOCK_FIRST_ROW + SKU_BLOCK_ROW_STEP * sku_idx
        sku_rows = extractions.loc[extractions[sku_col] == sku]
        markets_series = sku_rows["Market"].drop_duplicates()
        markets_list: list[str] = []
        for m in markets_series.tolist():
            if m is None or (isinstance(m, float) and pd.isna(m)):
                continue
            s = str(m).strip()
            if not s or s.lower() == "nan":
                continue
            markets_list.append(s)

        for market_offset, market in enumerate(markets_list):
            excel_row = base_row + market_offset
            if excel_row > SKU_BLOCK_LAST_ROW:
                log.warning(
                    "%s: SKU %r market rows exceed row %d; skipping remainder",
                    CANNIBAL_SHEET_NAME,
                    sku,
                    SKU_BLOCK_LAST_ROW,
                )
                break
            proj_row = proj_by_market.get(market)
            if proj_row is None:
                log.warning(
                    "%s: no extractions_proj row for Market %r (SKU %r row %d)",
                    CANNIBAL_SHEET_NAME,
                    market,
                    sku,
                    excel_row,
                )
                continue
            for j, col_idx in enumerate(year_col_indices):
                ws.cell(row=excel_row, column=col_idx, value=proj_row[CANNI_PERC_PROJ_COLS[j]])
            if ws_rev is not None:
                ws_rev.cell(
                    row=excel_row,
                    column=gp_rev_col_idx,
                    value=proj_row[CANNI_GP_PROJ_COL],
                )
            written += 1

    gp_target = f"{CANNIBAL_REV_SHEET_NAME}!{CANNI_GP_REV_COL_LETTER}" if ws_rev is not None else "(skipped)"
    log.info(
        "%s: wrote %d market row(s) (cols %s; GP → %s) from extractions_proj",
        CANNIBAL_SHEET_NAME,
        written,
        ",".join(CANNI_YEAR_COL_LETTERS),
        gp_target,
    )
    return written


def apply_insertions_to_workbook(
    wb: Workbook,
    extractions: pd.DataFrame,
    extractions_proj: pd.DataFrame,
) -> None:
    """Apply all template insertions for one source file to an open workbook."""
    if LAUNCH_YEAR_Q_COL not in extractions.columns:
        log.warning(
            "extractions has no %r column — Actual Launch Year/Month (K, L) will not be "
            "written on sku-block sheets. Reload extractions.csv or re-run extraction.",
            LAUNCH_YEAR_Q_COL,
        )
    else:
        n_launch = extractions[LAUNCH_YEAR_Q_COL].notna().sum()
        log.info(
            "extractions[%r]: %d non-null values (K=year, L=quarter start month)",
            LAUNCH_YEAR_Q_COL,
            n_launch,
        )

    sku_col = resolve_sku_column(extractions)
    unique_skus = list(extractions[sku_col].dropna().unique())
    if len(unique_skus) > MAX_SKU_SLOTS:
        log.warning(
            "%d unique SKUs exceed %d blocks; only the first %d will be written on sku-block sheets",
            len(unique_skus),
            MAX_SKU_SLOTS,
            MAX_SKU_SLOTS,
        )

    ws_home = wb["Home Tab"]
    ws_home.cell(
        row=HOME_TAB_BASE_CURRENCY_ROW,
        column=column_index_from_string(HOME_TAB_BASE_CURRENCY_COL),
        value=HOME_TAB_BASE_CURRENCY_VALUE,
    )
    log.info(
        "Home Tab: set base currency to %r at %s%s",
        HOME_TAB_BASE_CURRENCY_VALUE,
        HOME_TAB_BASE_CURRENCY_COL,
        HOME_TAB_BASE_CURRENCY_ROW,
    )

    for i, val in enumerate(unique_skus[:MAX_SKU_SLOTS]):
        ws_home.cell(row=6 + i, column=column_index_from_string("J"), value=val)

    write_sp_list_prices(wb["SP"], extractions)
    write_vol_auto_forecast_volumes(wb["VOL_Auto"], extractions)
    write_sales_ded_td_gtn(wb["SALES_DED_TD"], extractions)
    write_cogs_unit(wb["COGS"], extractions)
    write_cogs_other_combined(wb["COGS_OTHER"], extractions)
    write_sales_ded_ka_launch(wb["SALES_DED_KA"], extractions)

    write_ap_alloc_markets_column_x(wb[AP_ALLOC_SHEET_NAME], extractions)

    if extractions_proj is not None and not extractions_proj.empty:
        _require_columns(extractions_proj, ANP_PROJ_COLS, "extractions_proj")
        ws_ap = wb["A&P TABLE"]
        start_col = column_index_from_string("B")
        for i, (_, row_s) in enumerate(extractions_proj[ANP_PROJ_COLS].iterrows()):
            excel_row = 4 + i
            for j, col_name in enumerate(ANP_PROJ_COLS):
                ws_ap.cell(row=excel_row, column=start_col + j, value=row_s[col_name])

        # Unique Market (first-appearance order) → Home Tab column I from row 19 downward
        markets_u = extractions_proj["Market"].drop_duplicates()
        markets_u = [
            m for m in markets_u.tolist() if m is not None and not (isinstance(m, float) and pd.isna(m)) and str(m).strip()
        ]
        mi = column_index_from_string(HOME_TAB_UNIQUE_MARKET_COL)
        for off in range(HOME_TAB_UNIQUE_MARKET_CLEAR_ROWS):
            ws_home.cell(row=HOME_TAB_UNIQUE_MARKET_START_ROW + off, column=mi, value=None)
        for i, market in enumerate(markets_u):
            ws_home.cell(row=HOME_TAB_UNIQUE_MARKET_START_ROW + i, column=mi, value=market)
        log.info(
            "Home Tab: wrote %d unique Market value(s) to %s%s onward",
            len(markets_u),
            HOME_TAB_UNIQUE_MARKET_COL,
            HOME_TAB_UNIQUE_MARKET_START_ROW,
        )

        write_capex_sheet_ak_from_proj(wb[CAPEX_SHEET_NAME], extractions_proj)

    if extractions_proj is not None and not extractions_proj.empty:
        ws_rev = wb[CANNIBAL_REV_SHEET_NAME] if CANNIBAL_REV_SHEET_NAME in wb.sheetnames else None
        if ws_rev is None:
            log.warning("%s sheet missing — GP %% will not be written", CANNIBAL_REV_SHEET_NAME)
        write_cannibal_sheet_from_proj(
            wb[CANNIBAL_SHEET_NAME],
            extractions,
            extractions_proj,
            ws_rev=ws_rev,
        )


def npl_output_name(source_filename: str) -> str:
    """e.g. 'Project.xlsm' -> 'Project_NPL.xlsm'."""
    base, ext = os.path.splitext(source_filename)
    return f"{base}_NPL{ext}"


def copy_template_and_insert(
    template_path: str,
    output_path: str,
    extractions: pd.DataFrame,
    extractions_proj: pd.DataFrame,
) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path, keep_vba=True)
    try:
        apply_insertions_to_workbook(wb, extractions, extractions_proj)
        wb.save(output_path)
    finally:
        wb.close()
