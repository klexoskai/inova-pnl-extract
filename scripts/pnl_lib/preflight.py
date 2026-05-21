"""Pre-flight checks for old-format P&L → new-format insertion (no _NPL output)."""

from __future__ import annotations

import os
import shutil
import tempfile
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .excel_utils import basename_only, list_excel_files, resolve_sheet_name
from .extract_proj import extract_project_metrics
from .extract_sku import extract_sku_summary, preproc_sku_extractions
from .insert_template import (
    ANP_PROJ_COLS,
    CAPEX_PROJ_VALUE_COL,
    CANNI_GP_PROJ_COL,
    CANNI_PERC_PROJ_COLS,
    LAUNCH_YEAR_Q_COL,
    MAX_SKU_SLOTS,
    apply_insertions_to_workbook,
    resolve_sku_column,
)

DEFAULT_SHEET = "Project Summary"

READINESS_REPORT_COLUMNS = ["filename", "can_process", "notes"]

REQUIRED_TEMPLATE_SHEETS = [
    "Home Tab",
    "SP",
    "VOL_Auto",
    "SALES_DED_TD",
    "COGS",
    "COGS_OTHER",
    "SALES_DED_KA",
    "A&P ALLOC",
    "A&P TABLE",
    "CAPEX",
    "CANNIBAL",
    "CANNIBAL_REV",
]

SKU_VALUE_COLUMNS = [
    "list_price_AUD",
    "forecast_volume_y1",
    "GTN_perc excluding Launch yr one-time costs",
    "COGS/unit",
]


@dataclass
class FileReadiness:
    filename: str
    can_process: str
    notes: str
    file_path: str = ""
    issues: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def finalize(self) -> None:
        parts: list[str] = []
        if self.issues:
            parts.extend(self.issues)
        if self.warnings:
            parts.extend(self.warnings)
        self.notes = "; ".join(parts)
        self.can_process = "no" if self.issues else "yes"


def _check_template(template_path: str) -> list[str]:
    issues: list[str] = []
    if not os.path.isfile(template_path):
        issues.append(f"Template missing: {template_path}")
        return issues
    try:
        wb = load_workbook(template_path, read_only=True, keep_vba=True)
        try:
            missing = [s for s in REQUIRED_TEMPLATE_SHEETS if s not in wb.sheetnames]
            if missing:
                issues.append(f"Template missing sheets: {', '.join(missing)}")
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001
        issues.append(f"Template not readable: {type(exc).__name__}: {exc}")
    return issues


def _validate_sku_frame(df: pd.DataFrame, result: FileReadiness) -> None:
    if df.empty:
        result.issues.append("SKU extract: no rows (no active list-price markets)")
        return
    try:
        sku_col = resolve_sku_column(df)
    except KeyError:
        result.issues.append("SKU extract: no sku_name / sku_name.1 column")
        return
    n_skus = df[sku_col].dropna().nunique()
    n_rows = len(df)
    result.warnings.append(f"SKU extract: {n_rows} row(s), {n_skus} product SKU(s)")
    if n_skus > MAX_SKU_SLOTS:
        result.warnings.append(
            f"SKU count {n_skus} exceeds template limit {MAX_SKU_SLOTS} (only first {MAX_SKU_SLOTS} written)"
        )
    missing_cols = [c for c in SKU_VALUE_COLUMNS if c not in df.columns]
    if missing_cols:
        result.issues.append(f"SKU extract: missing columns {missing_cols!r}")
    if LAUNCH_YEAR_Q_COL not in df.columns:
        result.warnings.append(
            f"SKU extract: {LAUNCH_YEAR_Q_COL!r} missing — VOL_Auto K/L (launch year/month) will be blank"
        )
    elif df[LAUNCH_YEAR_Q_COL].notna().sum() == 0:
        result.warnings.append(
            f"SKU extract: {LAUNCH_YEAR_Q_COL!r} all empty — VOL_Auto K/L will be blank"
        )


def _validate_proj_frame(df: pd.DataFrame, result: FileReadiness) -> None:
    if df.empty:
        result.issues.append("Project extract: no rows (no active markets on Project Summary)")
        return
    missing_anp = [c for c in ANP_PROJ_COLS if c not in df.columns]
    if missing_anp:
        result.issues.append(f"Project extract: missing ANP columns {missing_anp!r}")
    n_markets = df["Market"].dropna().nunique() if "Market" in df.columns else 0
    result.warnings.append(f"Project extract: {len(df)} row(s), {n_markets} market(s)")
    if CAPEX_PROJ_VALUE_COL not in df.columns:
        result.warnings.append(f"Project extract: {CAPEX_PROJ_VALUE_COL!r} missing — CAPEX sheet AK will be skipped")
    cannibal_cols = [*CANNI_PERC_PROJ_COLS, CANNI_GP_PROJ_COL]
    missing_canni = [c for c in cannibal_cols if c not in df.columns]
    if missing_canni:
        result.warnings.append(
            f"Project extract: cannibal columns missing {missing_canni!r} — CANNIBAL sheet will be skipped"
        )


def _dry_run_insertion(
    template_path: str,
    sku_df: pd.DataFrame,
    proj_df: pd.DataFrame,
    result: FileReadiness,
) -> None:
    if result.issues:
        return
    fd, tmp_path = tempfile.mkstemp(suffix="_preflight.xlsm")
    os.close(fd)
    try:
        shutil.copy2(template_path, tmp_path)
        wb = load_workbook(tmp_path, keep_vba=True)
        try:
            apply_insertions_to_workbook(wb, sku_df, proj_df)
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001
        result.issues.append(f"Insertion dry-run failed: {type(exc).__name__}: {exc}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def assess_workbook(
    excel_path: str,
    *,
    template_path: str,
    apply_preproc: bool = True,
    run_insertion_dry_run: bool = True,
    template_issues: list[str] | None = None,
) -> FileReadiness:
    """
    Assess one old-format workbook for new-format insertion readiness.

    Does not write ``_NPL`` output.
    """
    filename = basename_only(excel_path)
    result = FileReadiness(filename=filename, can_process="no", notes="", file_path=excel_path)

    if template_issues:
        result.issues.extend(template_issues)

    if not os.path.isfile(excel_path):
        result.issues.append("File not found")
        result.finalize()
        return result

    try:
        resolve_sheet_name(excel_path, DEFAULT_SHEET)
    except Exception as exc:  # noqa: BLE001
        result.issues.append(f"Sheet {DEFAULT_SHEET!r} not found: {type(exc).__name__}: {exc}")
        result.finalize()
        return result

    sku_df = pd.DataFrame()
    proj_df = pd.DataFrame()

    try:
        sku_df = extract_sku_summary(excel_path, sheet_name=DEFAULT_SHEET)
        if apply_preproc and not sku_df.empty:
            sku_df = preproc_sku_extractions(sku_df)
    except Exception as exc:  # noqa: BLE001
        result.issues.append(f"SKU extract failed: {type(exc).__name__}: {exc}")

    try:
        proj_df = extract_project_metrics(excel_path, sheet_name=DEFAULT_SHEET)
    except Exception as exc:  # noqa: BLE001
        result.issues.append(f"Project extract failed: {type(exc).__name__}: {exc}")

    if not result.issues:
        _validate_sku_frame(sku_df, result)
    if not any(i.startswith("Project extract failed") for i in result.issues):
        _validate_proj_frame(proj_df, result)

    if run_insertion_dry_run and template_path and os.path.isfile(template_path):
        _dry_run_insertion(template_path, sku_df, proj_df, result)

    result.finalize()
    return result


def assess_directory(
    directory: str,
    *,
    template_path: str,
    apply_preproc: bool = True,
    run_insertion_dry_run: bool = True,
    progress: bool = True,
) -> list[FileReadiness]:
    """Run :func:`assess_workbook` on every Excel file under ``directory``."""
    directory = os.path.abspath(directory)
    paths = list_excel_files(directory)
    template_issues = _check_template(template_path)

    results: list[FileReadiness] = []
    iterator: Any = paths
    if progress:
        from tqdm import tqdm

        iterator = tqdm(paths, desc="Preflight", unit="file")

    for path in iterator:
        rel = os.path.relpath(path, directory)
        row = assess_workbook(
            path,
            template_path=template_path,
            apply_preproc=apply_preproc,
            run_insertion_dry_run=run_insertion_dry_run,
            template_issues=list(template_issues),
        )
        if rel != row.filename:
            row.warnings.insert(0, f"Path: {rel}")
        results.append(row)

    return results


def readiness_to_dataframe(results: list[FileReadiness]) -> pd.DataFrame:
    rows = [
        {
            "filename": r.filename,
            "can_process": r.can_process,
            "notes": r.notes,
        }
        for r in results
    ]
    return pd.DataFrame(rows, columns=READINESS_REPORT_COLUMNS)


def write_readiness_report_csv(results: list[FileReadiness], output_path: str) -> pd.DataFrame:
    df = readiness_to_dataframe(results)
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    df.to_csv(output_path, index=False)
    return df
